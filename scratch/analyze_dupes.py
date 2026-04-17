import openpyxl

master_xlsm_path = r"C:\Users\user\Documents\PROJECTS\UMA BALI CZ\MASTER FFE Frontend.xlsm"
wb = openpyxl.load_workbook(master_xlsm_path, data_only=True)
sheet = wb.active

from collections import defaultdict
codes_count = defaultdict(int)
categories = set()

for row in range(2, sheet.max_row + 1):
    code = sheet.cell(row=row, column=1).value
    cat = sheet.cell(row=row, column=2).value
    
    if code:
        codes_count[str(code).strip()] += 1
    if cat:
        categories.add(str(cat).strip().lower())

print(f"Total rows: {sheet.max_row}")
print(f"Distinct codes: {len(codes_count)}")
duplicates = {k:v for k,v in codes_count.items() if v > 1}
if duplicates:
    print(f"Duplicate codes found: {duplicates}")
else:
    print("No duplicate codes in XLSM.")

print(f"Categories: {categories}")
