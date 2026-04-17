import os
import glob
import json
import shutil
import openpyxl

codes = ['UMA-04', 'UMA-05', 'UMA-07', 'UMA-26', 'UMA-28', 'UMA-29', 'UMA-32', 'UMA-33', 'UMA-36', 'UMA-40', 'UMA-46', 'UMA-47', 'UMA-52', 'UMA-68', 'UMA-71', 'UMA-81']
brain_dir = r'C:\Users\user\.gemini\antigravity\brain\7c594f67-95aa-4eb8-94df-73d3c6e0db8e'
dest_dir = r'C:\Users\user\Documents\PROJECTS\UMA BALI CZ\frontend\public\images\products\catalog'
json_path = r'C:\Users\user\Documents\PROJECTS\UMA BALI CZ\frontend\src\data\products.json'
excel_path = r'C:\Users\user\Documents\PROJECTS\UMA BALI CZ\MASTER FFE Frontend.xlsm'

os.makedirs(dest_dir, exist_ok=True)

# 1. Map existing generated files
# They are named uma_04_premium_TIMESTAMP.png
gen_files = {}
for code in codes:
    prefix = code.lower().replace('-', '_') + '_premium'
    matches = glob.glob(os.path.join(brain_dir, f"{prefix}*.png"))
    if matches:
        gen_files[code] = max(matches, key=os.path.getctime)

# 2. Copy to frontend and update JSON
with open(json_path, 'r', encoding='utf-8') as f:
    products = json.load(f)

for code, src in gen_files.items():
    new_filename = f"{code.lower().replace('-', '_')}_premium.png"
    dest_path = os.path.join(dest_dir, new_filename)
    shutil.copy2(src, dest_path)
    # update json
    for p in products:
        if p['id'] == code:
            p['images'] = [f"/images/products/catalog/{new_filename}"]
            break

with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(products, f, indent=2, ensure_ascii=False)
print("Updated products.json")

# 3. Update Excel
wb = openpyxl.load_workbook(excel_path, keep_vba=True)
ws = wb.active

# Find columns
header_row = 1
code_col = None
img_col = None

for cell in ws[1]:
    if cell.value == 'item code':
        code_col = cell.column
    if cell.value == 'img file path':
        img_col = cell.column

if code_col and img_col:
    for row in range(2, ws.max_row + 1):
        c_val = ws.cell(row=row, column=code_col).value
        code_str = str(c_val).strip() if c_val else ''
        if code_str in gen_files:
            new_filename = f"{code_str.lower().replace('-', '_')}_premium.png"
            ws.cell(row=row, column=img_col).value = f"/images/products/catalog/{new_filename}"
    try:
        wb.save(excel_path)
        print("Updated Excel file.")
    except PermissionError:
        alt_path = excel_path.replace('.xlsm', '_updated.xlsm')
        wb.save(alt_path)
        print(f"File was locked. Saved instead to {alt_path}")
else:
    print("Could not find columns in Excel.")
