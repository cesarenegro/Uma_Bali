import openpyxl

master_xlsm_path = r"C:\Users\user\Documents\PROJECTS\UMA BALI CZ\MASTER FFE Frontend.xlsm"
wb = openpyxl.load_workbook(master_xlsm_path, data_only=True)
sheet = wb.active

print("Headers:")
for col in range(1, 15):
    val = sheet.cell(row=1, column=col).value
    print(f"Col {col}: {val}")
    
# Let's peek at row 2
print("\nRow 2:")
for col in range(1, 15):
    val = sheet.cell(row=2, column=col).value
    print(f"Col {col}: {val}")
