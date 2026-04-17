import openpyxl
import glob
import re
import os

master_xlsm_path = r"C:\Users\user\Documents\PROJECTS\UMA BALI CZ\MASTER FFE Frontend.xlsm"

def get_finish(materials):
    combined = " ".join(materials).lower()
    if 'wenge' in combined: return 'wenge'
    if 'grey' in combined or 'gray' in combined: return 'grey'
    return 'natural'

def clean_category(c):
    c = str(c).strip().lower()
    if c.endswith('s'): c = c[:-1]
    if c == 'sofa chair': c = 'chair'
    return c.replace(' ', '')

wb = openpyxl.load_workbook(master_xlsm_path, data_only=True)
sheet = wb.active

# Map code -> dict of info
code_db = {}
# Map current_db_path -> code
db_path_to_code = {}

for row in range(2, sheet.max_row + 1):
    code_raw = sheet.cell(row=row, column=1).value
    if not code_raw: continue
    code = str(code_raw).strip()
    
    cat_raw = sheet.cell(row=row, column=2).value
    cat = clean_category(cat_raw) if cat_raw else 'item'
    
    # materials
    mats = []
    for c in [7, 8, 9, 10]:
        val = sheet.cell(row=row, column=c).value
        if val: mats.append(str(val))
    
    finish = get_finish(mats)
    
    path_raw = sheet.cell(row=row, column=6).value
    if path_raw:
        path_norm = os.path.normpath(str(path_raw).strip())
        filename = os.path.basename(path_norm)
        db_path_to_code[filename] = code
        db_path_to_code[path_norm] = code

    code_db[code] = {
        'cat': cat,
        'finish': finish,
        'row': row
    }

# Now let's scan all physical files
image_dir = r"C:\Users\user\Documents\PROJECTS\UMA BALI CZ\frontend\public\images\products"
all_images = glob.glob(image_dir + '/**/*.png', recursive=True) + glob.glob(image_dir + '/**/*.jpg', recursive=True)

print(f"Total physical images: {len(all_images)}")

file_mappings = []

for img_path in all_images:
    name = os.path.basename(img_path)
    
    # Rule 1: does it match exactly something in DB?
    matched_code = db_path_to_code.get(name)
    if not matched_code:
        matched_code = db_path_to_code.get(img_path)
        
    # Rule 2: extract UMA-XX or uma_XX
    if not matched_code:
        m = re.search(r'(UMA|uma)[-_]?(\d{2})', name, re.IGNORECASE)
        if m:
            matched_code = f"UMA-{m.group(2)}"
            
    # Rule 3: extract number patterns if it's uniquely identifying
    if not matched_code:
        # check if there's exactly one 2-digit number that corresponds to a UMA code
        m = re.findall(r'\b(\d{2})\b', name)
        if m:
            potential_codes = [f"UMA-{c}" for c in m if f"UMA-{c}" in code_db]
            if len(potential_codes) == 1:
                matched_code = potential_codes[0]

    if matched_code and matched_code in code_db:
        file_mappings.append((img_path, matched_code))
    else:
        print(f"UNMATCHED FILE: {name}")

print(f"\nMatched {len(file_mappings)} files to codes.")

# Let's see mapping for one code as example
print("\nExample mapping for UMA-56:")
for f, c in file_mappings:
    if c == 'UMA-56':
        print(os.path.basename(f))
