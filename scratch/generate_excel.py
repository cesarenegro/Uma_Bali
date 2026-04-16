import json
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

base_dir = r"c:\Users\user\Documents\PROJECTS\UMA BALI CZ"
products_file = os.path.join(base_dir, "frontend", "src", "data", "products.json")
public_dir = os.path.join(base_dir, "frontend", "public")

with open(products_file, 'r', encoding='utf-8') as f:
    products = json.load(f)

# Required columns
headers = [
    "item code", "category", "subcategory", "description", "img", "img file path",
    "material", "upholstery material", "metal material", "marble type", 
    "dimension L", "dimension W", "dimension H", "weight", "volume", 
    "unit", "unit price (cost)", "multiplier", "selling price"
]

wb = Workbook()
ws = wb.active
ws.title = "MASTER FFE"

# Write headers
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Process rows
row_idx = 2
for item in products:
    item_code = item.get("id", "")
    category = item.get("category", "")
    subcategory = item.get("subcategory", "")
    description = item.get("description", "")
    
    # Image
    img_val = ""
    img_path = None
    img_file_path_str = ""
    if item.get("images") and len(item.get("images")) > 0:
        rel_img_path = item["images"][0].lstrip("/")
        abs_img_path = os.path.join(public_dir, rel_img_path.replace("/", "\\"))
        img_file_path_str = abs_img_path # Use full absolute path
        if os.path.exists(abs_img_path):
            img_path = abs_img_path
        else:
            img_val = rel_img_path
    
    material = ", ".join(item.get("materials", []))
    upholstery = item.get("upholstery", "")
    metal = item.get("metal", "")
    marble = item.get("marble", "")
    
    dims = item.get("dimensions", {})
    dim_L = dims.get("width", "")
    dim_W = dims.get("depth", "")
    dim_H = dims.get("height", "")
    
    weight = item.get("weight", "")
    volume = item.get("volume", "")
    unit = item.get("unit", "")
    
    unit_price = item.get("price", "")
    multiplier = item.get("multiplier", "")
    
    # Fill cells except selling price
    row_data = [
        item_code, category, subcategory, description, img_val, img_file_path_str,
        material, upholstery, metal, marble,
        dim_L, dim_W, dim_H, weight, volume,
        unit, unit_price, multiplier
    ]
    
    for col_idx, val in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=col_idx)
        c.value = val
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(col_idx == 4))
    
    # Selling Price Formula
    c = ws.cell(row=row_idx, column=len(headers))
    if unit_price != "" and multiplier != "":
        cost_col = get_column_letter(headers.index("unit price (cost)") + 1)
        mult_col = get_column_letter(headers.index("multiplier") + 1)
        c.value = f"={cost_col}{row_idx}*{mult_col}{row_idx}"
    c.alignment = Alignment(horizontal="center", vertical="center")
    
    # Process image insertion
    if img_path:
        try:
            img = Image(img_path)
            # Resize image to fit nicely
            # Max width roughly 100px, max height roughly 100px
            max_size = 100
            orig_width, orig_height = img.width, img.height
            ratio = min(max_size/orig_width, max_size/orig_height)
            new_width = int(orig_width * ratio)
            new_height = int(orig_height * ratio)
            img.width = new_width
            img.height = new_height
            
            img_col_letter = get_column_letter(headers.index("img") + 1)
            ws.add_image(img, f"{img_col_letter}{row_idx}")
            # Adjust row height specifically for image, 1 point ~ 1.33 pixels
            ws.row_dimensions[row_idx].height = new_height * 0.75 + 10
        except Exception as e:
            c = ws.cell(row=row_idx, column=headers.index("img") + 1)
            c.value = rel_img_path
            ws.row_dimensions[row_idx].height = 40
    else:
        # Give some default height for rows without images but with text
        ws.row_dimensions[row_idx].height = 40
        
    row_idx += 1

# Formatting
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border

# Column widths
column_widths = {
    "A": 12, "B": 15, "C": 15, "D": 40, "E": 18, "F": 40,
    "G": 20, "H": 20, "I": 15, "J": 15, 
    "K": 10, "L": 10, "M": 10, "N": 10, "O": 10, 
    "P": 10, "Q": 15, "R": 10, "S": 15
}
for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Freeze top row
ws.freeze_panes = "A2"

# Add AutoFilter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# Save
output_path = os.path.join(base_dir, "MASTER FFE Frontend.xlsx")
wb.save(output_path)
print(f"Generated {output_path}")
