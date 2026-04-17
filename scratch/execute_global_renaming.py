import openpyxl
import glob
import re
import os
import json
import shutil

# Paths
base_dir = r"C:\Users\user\Documents\PROJECTS\UMA BALI CZ"
master_xlsm_path = os.path.join(base_dir, "MASTER FFE Frontend.xlsm")
products_json_path = os.path.join(base_dir, "frontend", "src", "data", "products.json")
images_dir = os.path.join(base_dir, "frontend", "public", "images", "products")

# Helpers
def get_finish(materials):
    combined = " ".join(materials).lower()
    if 'wenge' in combined: return 'wenge'
    if 'grey' in combined or 'gray' in combined: return 'grey'
    return 'natural'

def clean_category(c):
    c = str(c).strip().lower()
    if c.endswith('s'): c = c[:-1]
    if c == 'sofa chair': c = 'chair'
    return c.replace(' ', '')

print("1. Parsing XLSM for product data...")
wb = openpyxl.load_workbook(master_xlsm_path, data_only=True) # use data_only first to read values
sheet = wb.active

# Map code -> dict of info
code_db = {}
db_path_to_code = {}

for row in range(2, sheet.max_row + 1):
    code_raw = sheet.cell(row=row, column=1).value
    if not code_raw: continue
    code = str(code_raw).strip()
    
    cat_raw = sheet.cell(row=row, column=2).value
    cat = clean_category(cat_raw) if cat_raw else 'item'
    
    mats = []
    for c in [7, 8, 9, 10]:
        val = sheet.cell(row=row, column=c).value
        if val: mats.append(str(val))
    
    finish = get_finish(mats)
    
    path_raw = sheet.cell(row=row, column=6).value
    if path_raw:
        path_norm = os.path.normpath(str(path_raw).strip())
        filename = os.path.basename(path_norm)
        db_path_to_code[filename.lower()] = code
        db_path_to_code[path_norm.lower()] = code

    code_db[code] = {
        'cat': cat,
        'finish': finish,
        'row': row,
        'code_clean': code.lower().replace('-', '') # e.g. UMA-18 -> uma18
    }

print("2. Finding and matching physical files...")
all_images = glob.glob(os.path.join(images_dir, '**', '*.png'), recursive=True) + \
             glob.glob(os.path.join(images_dir, '**', '*.jpg'), recursive=True)

# Mapped files structure: code -> list of absolute paths
mapped_files = {code: [] for code in code_db}

for img_path in all_images:
    name = os.path.basename(img_path)
    name_lower = name.lower()
    img_path_lower = os.path.normpath(img_path).lower()
    
    matched_code = db_path_to_code.get(name_lower) or db_path_to_code.get(img_path_lower)
        
    if not matched_code:
        m = re.search(r'(UMA|uma)[-_]?(\d{1,2})', name, re.IGNORECASE)
        if m:
            matched_code = f"UMA-{int(m.group(2)):02d}"
            
    if not matched_code:
        m = re.findall(r'\b(\d{2})\b', name)
        if m:
            potential_codes = [f"UMA-{c}" for c in m if f"UMA-{c}" in code_db]
            if len(potential_codes) == 1:
                matched_code = potential_codes[0]

    if matched_code and matched_code in code_db:
        # Avoid dupes if same file matched twice
        if img_path not in mapped_files[matched_code]:
            mapped_files[matched_code].append(img_path)

suffix_sequence = ['', '_A', '_B', '_C', '_D', '_E', '_F', '_G']

print("3. Executing renaming and gathering new paths...")

products_new_dir = os.path.join(base_dir, "frontend", "public", "images", "products_new")
if os.path.exists(products_new_dir):
    shutil.rmtree(products_new_dir)
os.makedirs(products_new_dir)

code_to_new_images = {} # code -> list of relative paths for JSON

for code, files in mapped_files.items():
    if not files:
        continue
    
    # Sort files so we have consistent naming. Let's make sure 'premium' files come first if they exist!
    def sort_key(f):
        is_premium = 0 if 'premium' in f.lower() else 1
        return (is_premium, f.lower())

    files.sort(key=sort_key)
    
    cat = code_db[code]['cat']
    c_clean = code_db[code]['code_clean']
    finish = code_db[code]['finish']
    
    new_relative_paths = []
    
    for i, old_path in enumerate(files):
        suffix = suffix_sequence[i] if i < len(suffix_sequence) else f"_{i}"
        ext = os.path.splitext(old_path)[1].lower()
        new_name = f"{cat}_{c_clean}_{finish}{suffix}{ext}"
        new_path = os.path.join(products_new_dir, new_name)
        
        # Copy to staging directory
        if not os.path.exists(new_path):
            shutil.copy2(old_path, new_path)
            
        rel_target = f"/images/products/{new_name}"
        new_relative_paths.append(rel_target)
        
    code_to_new_images[code] = new_relative_paths

print("3b. Swapping directories to complete rename...")
# Remove old directory
shutil.rmtree(images_dir)
# Rename the staging directory to the actual directory name
os.rename(products_new_dir, images_dir)

print("4. Updating JSON...")
with open(products_json_path, 'r', encoding='utf-8') as f:
    products = json.load(f)

for p in products:
    code = p['id']
    if code in code_to_new_images and code_to_new_images[code]:
        p['images'] = code_to_new_images[code]

with open(products_json_path, 'w', encoding='utf-8') as f:
    json.dump(products, f, indent=2, ensure_ascii=False)

print("5. Updating XLSM (preserving VBA)...")
wb_vba = openpyxl.load_workbook(master_xlsm_path, keep_vba=True)
ws = wb_vba.active

code_col, img_col = None, None
for cell in ws[1]:
    if cell.value == 'item code': code_col = cell.column
    if cell.value == 'img file path': img_col = cell.column

if code_col and img_col:
    for row in range(2, ws.max_row + 1):
        c_val = ws.cell(row=row, column=code_col).value
        code = str(c_val).strip() if c_val else ''
        if code in code_to_new_images and code_to_new_images[code]:
            first_image = code_to_new_images[code][0]
            ws.cell(row=row, column=img_col).value = first_image
            
    try:
        wb_vba.save(master_xlsm_path)
        print("Updated Excel file.")
    except PermissionError:
        alt_path = master_xlsm_path.replace('.xlsm', '_updated.xlsm')
        wb_vba.save(alt_path)
        print(f"File was locked. Saved instead to {alt_path}")
else:
    print("Could not find columns in Excel.")

print("All tasks completed successfully!")
